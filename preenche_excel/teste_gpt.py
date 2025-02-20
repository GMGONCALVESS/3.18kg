import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine, text
from openpyxl import load_workbook

# Database connection
engine = create_engine('postgresql://postgres:admin@192.168.88.61:5432/yield_debentures')

# Set reference date
data_atual = "2025-02-11"

# Load the Excel file
file_path = "copia_teste.xlsx"
lista = pd.read_excel(file_path, sheet_name=1)

# Extract the specific rows and columns to modify
modificar = lista[["CodCustodia", "Duration", "VlMrc"]][160:193]

# Fetch duration values from the database
pega_duration = pd.read_sql(f"SELECT DISTINCT codigo_ativo, duration FROM dados_debenture WHERE data_referencia = '{data_atual}'", engine)

# Merge the DataFrames on 'CodCustodia' and 'codigo_ativo'
merged = modificar.merge(pega_duration, left_on="CodCustodia", right_on="codigo_ativo", how="left")

# Drop duplicate columns
merged.drop(columns=["codigo_ativo", "Duration"], inplace=True)

# Update "Duration" in modificar using the merged data
for ind, row in merged.iterrows():
    if pd.notna(row.array[2]):  # Avoid issues with NaN values
        modificar.loc[modificar.index[ind], "Duration"] = row.array[2]

# **Step 2: Save `modificar` Back to Excel Without Overwriting Other Data**
wb = load_workbook(file_path)
ws = wb["Planilha1"]  # Ensure you are using the correct sheet name

# Loop through modificar and update only the "Duration" column in the existing Excel sheet
for index, row in modificar.iterrows():
    excel_row = index + 2  # Adjust for Excel (1-based index)
    ws[f"U{excel_row}"] = row["Duration"]  # Assuming "Duration" is in column U

# Save the updated file
wb.save(file_path)

print(modificar)

soma1 = soma2 = 0
for ind, row in  modificar.iterrows():
    if pd.notna(row.array[2]) and pd.notna(row.array[1]):
        soma1 = row.array[1]*row.array[2]
        soma2 = row.array[2]

duration_media = soma1/soma2

print(duration_media, duration_media/252)
